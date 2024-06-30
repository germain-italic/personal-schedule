import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta

def create_schedule():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Planning Personnel"

    # Définir les couleurs pour chaque activité (tons pastels)
    colors = {
        "Petit déjeuner": "FFE5B4",
        "Salle de sport": "E6FFE6",
        "Douche": "FFD1DC",
        "Travail": "FFB3BA",
        "Courses": "FFCCCC",
        "Déjeuner": "D7BDE2",
        "Dîner": "D5D8DC",
        "Sommeil": "B3E0FF",
        "Ménage": "C5E1A5",
        "Cuisine": "F9E79F",
        "Réunion familiale": "ABEBC6",
        "Réunion amis": "FAD7A0",
        "Temps libre": "FFFFFF"
    }

    # Créer l'en-tête
    days = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    for col, day in enumerate(days, start=2):
        sheet.cell(row=1, column=col, value=day)
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 17

    # Créer les créneaux horaires
    start_time = datetime.strptime("00:00", "%H:%M")
    for row in range(2, 50):  # 24 heures * 2 (30 min intervals)
        time = start_time + timedelta(minutes=30 * (row - 2))
        sheet.cell(row=row, column=1, value=time.strftime("%H:%M"))
    sheet.column_dimensions['A'].width = 17

    # Structure du tableau activities :
    # Chaque élément est un tuple de 4 éléments :
    # ("Nom de l'activité", durée, jours, "heure de début")
    #
    # 1. Nom de l'activité : string
    # 2. Durée : integer (en demi-heures, ex: 2 = 1 heure, 4 = 2 heures)
    # 3. Jours : objet range ou liste
    #    - range(5) : [0, 1, 2, 3, 4] (lundi à vendredi)
    #    - range(7) : [0, 1, 2, 3, 4, 5, 6] (tous les jours)
    #    - [5, 6] : samedi et dimanche
    #    - [0, 2, 4] : lundi, mercredi, vendredi
    # 4. Heure de début : string au format "HH:MM"
    #
    # Exemples :
    # ("Sommeil", 18, range(5), "22:00")  # 9h de sommeil, lun-ven, début à 22h
    # ("Weekend", 24, [5, 6], "22:00")    # 12h de sommeil, sam-dim, début à 22h
    # ("Travail", 16, range(5), "09:00")  # 8h de travail, lun-ven, début à 9h
    # ("Sport", 4, [1, 3, 5], "18:00")    # 2h de sport, mar-jeu-sam, début à 18h
    #
    # Pour modifier :
    # - Ajoutez, supprimez ou modifiez les tuples dans la liste activities
    # - Assurez-vous que les activités ne se chevauchent pas dans le temps
    # - Vérifiez que la durée totale des activités par jour ne dépasse pas 48 (24h)

    # Remplir le planning
    activities = [
        ("Sommeil", 18, range(5), "22:00"),
        ("Sommeil", 23, [5, 6], "22:00"),
        ("Petit déjeuner", 1, range(5), "07:00"),
        ("Petit déjeuner", 1, [5, 6], "09:30"),
        ("Salle de sport", 4, range(5), "07:30"),
        ("Douche", 1, range(5), "09:30"),
        ("Courses", 3, [1, 4], "10:00"),
        ("Travail", 16, range(5), "11:00"),
        ("Déjeuner", 2, range(5), "12:30"),
        ("Déjeuner", 2, [5, 6], "12:30"),
        ("Dîner", 2, range(7), "19:00"),
        ("Ménage", 4, [5], "10:00"),
        ("Douche", 1, [5, 6], "12:00"),
        ("Cuisine", 4, [6], "10:00"),
        ("Réunion familiale", 3, [6], "20:00"),
        ("Réunion amis", 6, [4], "20:00")
    ]

    # Initialiser toutes les cellules comme "Temps libre"
    for row in range(2, 50):
        for col in range(2, 9):
            cell = sheet.cell(row=row, column=col)
            cell.value = "Temps libre"
            cell.fill = PatternFill(start_color=colors["Temps libre"], end_color=colors["Temps libre"], fill_type="solid")

    # Fonction pour obtenir l'index de ligne à partir de l'heure
    def get_row_index(time_str):
        time = datetime.strptime(time_str, "%H:%M")
        return (time.hour * 2) + (time.minute // 30) + 2

    # Remplir le planning avec les activités
    total_time = {activity: 0 for activity in colors.keys()}
    for activity, duration, days, start_time in activities:
        start_row = get_row_index(start_time)
        color = colors[activity]
        for day in days:
            for i in range(duration):
                row = (start_row + i - 1) % 48 + 2  # Wrap around to the next day if necessary
                cell = sheet.cell(row=row, column=day + 2)
                cell.value = activity
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                total_time[activity] += 30  # Ajouter 30 minutes à chaque fois

    # Calculer le temps libre total
    total_time_activities = sum(total_time.values())
    total_time["Temps libre"] = 7 * 24 * 60 - total_time_activities

    # Ajouter la liste des activités avec leurs fréquences, contraintes et temps total
    start_col = 10
    sheet.cell(row=1, column=start_col, value="Activités, fréquences, contraintes et temps total:")
    activity_details = [
        ("Petit déjeuner", "Quotidien, matin, 30 min (7h en semaine, 10h le weekend)"),
        ("Salle de sport", "5 fois par semaine, matin, 2h"),
        ("Douche", "Quotidienne, matin, 30 min (à la salle de sport ou à la maison)"),
        ("Travail", "2 x 4h quotidiennes du lundi au vendredi, commence à 11h"),
        ("Courses", "2 fois par semaine, 1h30, matin avant le travail"),
        ("Déjeuner", "Quotidien, 1h, commence à 12h30 (14h le samedi, 13h30 le dimanche)"),
        ("Dîner", "Quotidien, 1h"),
        ("Sommeil", "9h par nuit (jusqu'à 10h le weekend)"),
        ("Ménage", "3h / semaine, samedi matin avant le Déjeuner"),
        ("Cuisine", "4h / semaine, dimanche matin avant le Déjeuner"),
        ("Réunion familiale", "1h30 / semaine, dimanche à 20h"),
        ("Réunion amis", "3h / semaine, vendredi soir"),
        ("Temps libre", "Plages de temps non affectées")
    ]
    for row, (activity, details) in enumerate(activity_details, start=2):
        sheet.cell(row=row, column=start_col).fill = PatternFill(start_color=colors[activity], end_color=colors[activity], fill_type="solid")
        sheet.cell(row=row, column=start_col+1, value=activity)
        sheet.cell(row=row, column=start_col+2, value=details)
        if activity in total_time:
            hours, minutes = divmod(total_time[activity], 60)
            sheet.cell(row=row, column=start_col+3, value=f"{hours}h{minutes:02d}")

    # Ajuster la largeur des colonnes
    for col in range(start_col, start_col+4):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30

    # Sauvegarder le fichier
    wb.save("planning_personnel.xlsx")

create_schedule()